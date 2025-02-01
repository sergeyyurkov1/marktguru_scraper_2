import random
import string
import time
import warnings
from datetime import date

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

from helpers import rank_similarity


def set_location(driver, first_item: str, zip_: str):
    driver.get(
        f"https://www.marktguru.de/search/{first_item}?title={first_item}&page=0"
    )

    time.sleep(15)  # Waits for the page to load

    # Tries to remove view-blocking elements
    e = driver.find_element(By.ID, "usercentrics-root")
    driver.execute_script("arguments[0].remove();", e)

    # Opens location input
    location_input = driver.find_element(By.CLASS_NAME, "location-default-text")
    # print(location_input)
    location_input.click()

    time.sleep(5)

    # Selects the location input and enters the ZIP code
    inputs = driver.find_elements(By.TAG_NAME, "input")
    inputs[1].send_keys(zip_ + Keys.ENTER)

    time.sleep(5)

    # Selects the first address
    driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
    time.sleep(2)
    driver.switch_to.active_element.send_keys(Keys.ENTER)

    time.sleep(5)


def search_page(
    driver, url: str, item: str, page: int, zip_: str, store_data: dict
) -> list:
    results = []

    driver.get(f"{url}/{item}?title={item}&page={page}")

    # Exit condition: last page found, continues with the next item
    # ---------------------------
    WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.CLASS_NAME, "headline"))
    )
    assert item.upper() in driver.find_element(By.CLASS_NAME, "headline").text

    # Waits until the Item cards are loaded - affects the results
    WebDriverWait(driver, 120).until(
        EC.visibility_of_all_elements_located((By.TAG_NAME, "li"))
    )

    # ---------------------------
    location_text = driver.find_element(By.CLASS_NAME, "location-text").text
    # print(location_text)
    if zip_ not in location_text:
        # print("Warning: Location not set")
        ...

    # Parses the page
    # ---------------------------
    bs = BeautifulSoup(driver.page_source, "html.parser")

    lis = bs.select("li")
    not_found = 0
    for li in lis:
        name = li.select(store_data["name"])
        i = {}
        if len(name) > 0:
            # Item
            # ---------------------------
            i["Item"] = item

            # Name
            # ---------------------------
            i["Name"] = "".join([x.text for x in name]).rstrip().lower()

            # Date valid
            # ---------------------------
            dd = li.select(store_data["dv"])
            i["Date valid"] = "".join([x.text for x in dd]).rstrip().lower()

            # Store
            # ---------------------------
            a = li.select(store_data["store1"] + " > a")
            if len(list(a)) == 0:
                store = li.select(store_data["store1"])
                if len(list(store)) == 0:
                    span = li.select(store_data["store1"] + " > span")
                    i["Store"] = "".join([x.text for x in span]).rstrip().lower()
                else:
                    i["Store"] = "".join([x.text for x in store]).rstrip().lower()
            else:
                i["Store"] = "".join([x.text for x in a]).rstrip().lower()

            # Brand
            # ---------------------------
            brand_a = li.select(store_data["brand1"] + " > a")
            if len(list(brand_a)) == 0:
                brand = li.select(store_data["brand1"])
                if len(list(brand)) == 0:
                    brand_span = li.select(store_data["brand1"] + " > span")
                    i["Brand"] = "".join([x.text for x in brand_span]).rstrip().lower()
                else:
                    i["Brand"] = "".join([x.text for x in brand]).rstrip().lower()
            else:
                i["Brand"] = "".join([x.text for x in brand_a]).rstrip().lower()

            # Price
            # ---------------------------
            price_strong = li.select(store_data["price1"])
            if len(list(price_strong)) == 0:
                price_dd = li.select(store_data["price2"])
                price_p = li.select(store_data["price3"])

                i["Price"] = "".join([x.text for x in price_dd]).rstrip().lower()
                i["Note"] = "".join([x.text for x in price_p]).rstrip().lower()
            else:
                i["Price"] = (
                    "".join([x.text for x in price_strong])
                    .split("-")[0]
                    .rstrip()
                    .lower()
                )

            # print(i)  # for debugging

            results.append(i)
        else:
            not_found += 1
            continue

    # For debugging
    # with open("debug.html", "w") as f:
    #     f.write(driver.page_source)
    # with open("results.txt", "w") as f:
    #     f.write(repr(results))

    if len(lis) == not_found:
        raise Exception(
            "Warning: Getting empty results. Please check the 'Name' selector for changes, save the settings, and retry"
        )

    return results


def launch_scraper(driver, url, shopping_list, zip_, store_data, set_progress):
    data = []
    for item in shopping_list:
        page = 0
        while True:
            set_progress(("Scraping", ": ", f"'{item}' - page {page + 1}", 60))

            try:
                page_results = search_page(driver, url, item, page, zip_, store_data)

                data.extend(page_results)
            except AssertionError:
                # Reached the last page
                break
            except Exception:
                raise

            page += 1

    return data


def generate_output(data: list, lp: str, item_blacklist: list, filter: bool) -> str:
    warnings.simplefilter(action="ignore", category=FutureWarning)

    df_len = len(data)

    df = pd.DataFrame(data)

    # Checks empty columns and warns about selector issues
    # ---------------------------
    empty_columns = []
    for column in df.columns:
        if (df[column].values == "").sum() == df_len:
            empty_columns.append(column)
    if len(empty_columns) > 0:
        empty_columns_str = ", ".join([f"'{x}'" for x in empty_columns])
        raise Exception(
            f"Warning: {empty_columns_str} column(s) empty. Please check HTML tags for changes, save the settings, and retry"
        )

    # Blacklist
    # ---------------------------
    df = df[~df["Name"].isin(item_blacklist)]

    # Auto filter
    if filter:
        df = rank_similarity(df)

    # Handles Price and Units
    # ---------------------------
    try:
        df[["Price", "Unit"]] = df["Price"].str.split("/", expand=True)
    except:
        raise Exception(
            "Warning: Issue with the 'Price'/'Unit' selector. Please check HTML tags for changes, save the settings, and retry"
        )

    def str_to_float(x: str) -> float:
        try:
            return float(x.split(" ")[1].replace(",", "."))
        except ValueError:
            return 999.9  # returns 999.9 if cannot process the price

    try:
        df["Price"] = df["Price"].apply(lambda x: str_to_float(x))
        df.sort_values(
            ["Store", "Item", "Price"], ascending=True, inplace=True
        )  # sorting the Price from lowest to highest
    except:
        raise Exception(
            "Warning: Issue with the fallback 'Price' selector. Please check HTML tags for changes, save the settings, and retry"
        )

    # Reordering columns
    # ---------------------------
    try:
        df = df[
            ["Store", "Item", "Name", "Brand", "Price", "Unit", "Date valid", "Note"]
        ]
    except KeyError:
        df = df[["Store", "Item", "Name", "Brand", "Price", "Unit", "Date valid"]]

    # Lowest price indicator
    # ---------------------------
    df["LP"] = df.groupby([lp])["Price"].transform("min")
    df["Lowest price across stores"] = df.apply(
        lambda x: f"✅ {x[lp]}" if x["LP"] == x["Price"] else "", axis=1
    )
    df.drop(["LP"], axis=1, inplace=True)  # remove the temporary column

    # Removing duplicate entries
    # ---------------------------
    df.drop_duplicates(
        subset=["Name", "Brand", "Price", "Unit", "Date valid"],
        keep=False,
        inplace=True,
    )

    # Note selector check
    # ---------------------------
    # TODO

    # Handles the output
    # ------------------------------------------------------
    today = date.today()
    postfix = "".join(
        random.choices(string.ascii_uppercase + string.ascii_lowercase, k=5)
    )

    gbo = df.groupby(["Store"])

    # Creates an empty Excel file
    blank_line = pd.DataFrame()
    blank_line.to_excel(f"{today}_{postfix}.xlsx", index=False)

    # Loads the file, sets up ExcelWriter
    book = load_workbook(f"{today}_{postfix}.xlsx")
    writer = pd.ExcelWriter(f"{today}_{postfix}.xlsx", engine="openpyxl")
    writer.sheets.update({ws.title: ws for ws in book.worksheets})

    # Applies formatting to grouped by tables and writes to the file
    for e, (_, o) in enumerate(gbo):
        o[["Store", "Item"]] = o[["Store", "Item"]].where(
            o[["Store", "Item"]].apply(lambda x: x != x.shift()), ""
        )
        if e == 0:
            o.to_excel(writer, index=False)
        else:
            o.to_excel(
                writer, startrow=writer.sheets["Sheet1"].max_row + 1, index=False
            )

    writer.close()

    return f"{today}_{postfix}.xlsx"
